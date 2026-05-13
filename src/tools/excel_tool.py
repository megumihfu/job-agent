from crewai.tools import BaseTool
import pandas as pd
from datetime import datetime
import os
import json
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

class ExcelExportTool(BaseTool):
    name: str = "Excel export tool"
    description: str = """
        Exports filtered job offers to an Excel file.
        Use this tool at the end of the analysis to save only the kept offers.
    """
    
    def _run(self, jobs_data: str) -> str:

        os.makedirs("outputs", exist_ok=True)
        
        try:
            jobs = json.loads(jobs_data)
        except:
            with open("outputs/empty_run.txt", "w") as f:
                f.write("No jobs found or invalid data")
            return "please pass job data as JSON"
        
        if not jobs:
            return "No jobs to export"
        
        df = pd.DataFrame(jobs)
        df.columns = [c.lower() for c in df.columns] 
        df = df.loc[:, ~df.columns.duplicated()]
        df = df.rename(columns={
                'position': 'title',
                'jobUrl': 'url'
            })

        if 'status' not in df.columns:
            df.insert(0, 'status', 'Not applied')
        
        required_columns = ['title', 'company', 'location', 'country', 'salary', 'url']
        for col in required_columns:
            if col not in df.columns:
                df[col] = 'N/A'
            else:
                df[col] = df[col].fillna('N/A').replace('', 'N/A')

        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"outputs/job_offers_{timestamp}.xlsx"
        
        df.to_excel(filename, index=False, sheet_name='Job offers')
        self._apply_excel_formatting(filename)

        return f"Excel file generated: {filename} with {len(jobs)} job offers"
    
    def _apply_excel_formatting(self, filepath):
        """ apply conditional formatting & styling to the Excel file"""
        red_fill = PatternFill(start_color='FF6B6B', end_color='FF6B6B', fill_type='solid')  
        green_fill = PatternFill(start_color='85E085', end_color='85E085', fill_type='solid')  
        yellow_fill = PatternFill(start_color='FFE066', end_color='FFE066', fill_type='solid') 
        header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        wb = load_workbook(filepath)
        ws = wb.active
        
        status_col_idx = None
        for cell in ws[1]:
            if cell.value and str(cell.value).lower() == 'status':
                status_col_idx = cell.column
                status_col_letter = cell.column_letter 
                break
        
        if not status_col_idx:
            status_col_idx = 1
            status_col_letter = 'A'
        
        dv = DataValidation(type="list", formula1='"Not applied, Applied, In progress, Interview, Rejected"')
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        status_values = {
            'Not applied': red_fill,
            'Applied': green_fill,
            'In progress': yellow_fill,
            'Interview': yellow_fill,
            'Rejected': red_fill
        }
        
        for row_idx in range(2, ws.max_row + 1):
            status_cell = ws[f'{status_col_letter}{row_idx}']
            status_value = status_cell.value
            
            for key, fill_color in status_values.items():
                if status_value and key in str(status_value):
                    status_cell.fill = fill_color
                    break
            
            dv.add(status_cell) 
        
        ws.add_data_validation(dv)
        
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50) 
            ws.column_dimensions[col_letter].width = adjusted_width
        
        ws.freeze_panes = ws['A2']
        wb.save(filepath)